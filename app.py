from flask import Flask, make_response
from flask import request
import openpyxl
import json
import requests

MAX_HEADER_ROWS_COUNT = 4
VERIFICATION_TARGET_CELL_INDEX = 0
DESIGN_VALUE_CELL_INDEX = 3
NEED_VERIFY_CELL_INDEX = 4

app = Flask(__name__)


@app.route("/")
def hello_world():
    return "hello, world!"


@app.route("/expert/review", methods=['POST'])
def review():
    excel_file = request.files['file']

    # URL 쿼리 파라미터로 검증대상 타입과 대상 partNo 수신.
    # 검증 대상. ex) "IC", "Transistor", "Diode"
    target_type = request.args.get('validTarget')

    # 플랫폼 백엔드에서 생성한 세션 ID
    session_id = request.args.get('sessionId')

    # 검증 대상 partNo. ex) "LM2576HVSX-ADJ/NOPB"
    # target_part_no = request.args.get('partNo')

    # 엑셀 파일 받아서 파싱
    # 검증 필요한 열들 플랫폼에 검증 요청
    header_rows, parsed_rows = __parse_excel(excel_file)

    # DB에 조회할 검증 대상 번호 ex) 'LM2576HVSX-ADJ/NOPB'
    target_part_no = header_rows[0][3]

    # 플랫폼에 최초 1회 호출하는 API
    # 검증 대상 조회 세션에 문제가 생길 시 최종 response 값으로 {"review_start": false} 전달
    start_check = __review_start_from_platform(valid_target=target_type, target_part_no=target_part_no,
                                               session_id=session_id)

    # Todo: refactor
    if not start_check['start_check']:
        # response = {"review start": False} # key에 공백이 들어가면 안됩니다.
        response = make_response(json.dumps({"review_start": False}))
        return response

    # 플랫폼에서 검증된 열들의 결과값 저장
    review_results = []
    for row in parsed_rows:
        verification_target = row['partName']  # 검증 항목 cell
        design_value = row['designValue']  # 설계 값 cell
        print(row['partName'])
        print(design_value)
        validate_result = __request_part_to_platform(part_no=target_part_no, verification_target=verification_target,
                                                     design_value=design_value, session_id=session_id)
        if validate_result['valid'] is not None:
            review_results.append(validate_result['valid'])

    # Todo: refactor
    # 검증된 열들의 결과값을 종합 처리후 응답
    passvalidate_results = []
    verification_result = True
    for verification, excel_row in zip(review_results, parsed_rows):

        if not verification:
            verification_result = False
        pass_validate_result = {
            "partName": excel_row["partName"],
            "designValue": excel_row["designValue"],
            "verification": verification,
        }
        passvalidate_results.append(pass_validate_result)

    response = {
        # "partNo": "LM2576HVSX-ADJ/NOPB"
        header_rows[0][VERIFICATION_TARGET_CELL_INDEX]: header_rows[0][DESIGN_VALUE_CELL_INDEX],
        # header_rows[1][0]: header_rows[1][DESIGN_VALUE_CELL_INDEX], # "type": "step_down" ### 엑셀 항목이 비어있어서 추후에 처리
        # "manufacturer_name": "TI"
        header_rows[2][VERIFICATION_TARGET_CELL_INDEX]: header_rows[2][DESIGN_VALUE_CELL_INDEX],
        "passReview": verification_result,
        "reviewResults": passvalidate_results
    }

    # response = review_results
    # 응답 예시
    # return response = {
    #     "partNo": "LM2576HVSX-ADJ/NOPB",
    #     "type": "step_down",
    #     "manufacturer_name": "TI",
    #     "passReview": false,
    #     "reviewResults": [
    #         {
    #             "partName": "oprating_temperature_min",
    #             "designValue": -20,
    #             "verification": true,
    #         },
    #         {
    #             "partName": "oprating_temperature_max",
    #             "designValue": 100,
    #             "verification": false,
    #         },
    #         {
    #             "partName": "storage_temperature_min",
    #             "designValue": -45,
    #             "verification": true,
    #         },
    #         {
    #             "partName": "storage_temperature_max",
    #             "designValue": 110,
    #             "verification": false,
    #         }
    #     ]
    # }
    print(response)

    response = make_response(json.dumps(response))
    response.headers['Content-Type'] = 'application/json'
    return response


def __parse_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[0]]

    # Header rows 처리. 검증 항목, 비고, 단위, 설계 값, 적용여부. 1행 한글 row 제외
    # partNo, type, manufactureer_name. header_rows 넣음
    header_rows = []
    header_rows_range = sheet.iter_rows(min_row=2, max_row=MAX_HEADER_ROWS_COUNT, values_only=True)
    for row in header_rows_range:
        header_rows.append(row)

    # Header rows 제외
    # 적용 여부 체크된 row만 parsed_rows 넣음
    parsed_rows = []
    rows_range = sheet.iter_rows(min_row=MAX_HEADER_ROWS_COUNT, max_row=sheet.max_row, values_only=True)
    for row in rows_range:
        # 빈 row 무시
        if row is None:
            continue

        # 적용여부 체크 안 되어있으면 무시
        if row[NEED_VERIFY_CELL_INDEX] != 'o':
            continue

        parsed_row = {'partName': row[VERIFICATION_TARGET_CELL_INDEX], 'designValue': row[DESIGN_VALUE_CELL_INDEX]}
        parsed_rows.append(parsed_row)

    return header_rows, parsed_rows


def __review_start_from_platform(valid_target, target_part_no, session_id):
    data = {'validTarget': valid_target, 'partNo': target_part_no}

    # dockerized url
    # url = "http://platform/review/start"

    # local url
    url = "http://localhost:8080/review/start"
    # url = "http://localhost:80/review/start"
    headers = {'Content-Type': 'application/json'}
    body = json.dumps(data)
    params = {'sessionId': session_id}  # URL 쿼리 파라미터

    # response 예시
    # {'start_check' : true}
    print("review start")
    print(body)
    response = requests.post(url, data=body, headers=headers, params=params).json()
    return response


def __request_part_to_platform(part_no, verification_target, design_value, session_id):
    # dockerized url
    # url = "http://platform/review/part"

    # local url
    url = "http://localhost:8080/review/part"
    # url = "http://localhost:80/review/part"
    headers = {'Content-Type': 'application/json'}

    # body = parsed_row
    body = json.dumps({
        "partNo": part_no,
        "verificationTarget": verification_target,
        "designValue": design_value
    })
    params = {'sessionId': session_id}  # URL 쿼리 파라미터

    # response 예시
    # {'verificationTarget': 'oprating_temperature_max', 'valid': true'}
    response = requests.post(url, data=body, headers=headers, params=params).json()
    print(response)
    return response


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=False, port=5000)
